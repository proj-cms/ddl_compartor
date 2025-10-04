import unittest
import pandas as pd
import logging
from src.db_compare import compare_ddls
from src.excel.excel_writer import ExcelWriter

logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s %(levelname)s %(name)s %(message)s')

class TestDDLComparator(unittest.TestCase):
    def test_compare_ddls(self):
        import openpyxl
        try:
            diff, only_db1, only_db2, result_path = compare_ddls('src/config.yaml')
            self.assertIsInstance(diff, pd.DataFrame)
            self.assertIsInstance(only_db1, pd.DataFrame)
            self.assertIsInstance(only_db2, pd.DataFrame)
            # Check Excel file
            import os
            self.assertTrue(os.path.exists(result_path), f"Result Excel file not found: {result_path}")
            wb = openpyxl.load_workbook(result_path)
            self.assertIn('DiffColumns', wb.sheetnames)
            self.assertIn('OnlyInDB1', wb.sheetnames)
            self.assertIn('OnlyInDB2', wb.sheetnames)
            # Check DiffColumns sheet has table_name and column_name columns
            diff_ws = wb['DiffColumns']
            diff_header = [cell.value for cell in diff_ws[1]]
            self.assertIn('table_name', diff_header)
            self.assertIn('column_name', diff_header)
            # Check OnlyInDB1 sheet has table_name and column_name columns
            db1_ws = wb['OnlyInDB1']
            db1_header = [cell.value for cell in db1_ws[1]]
            self.assertIn('table_name', db1_header)
            self.assertIn('column_name', db1_header)
            # Check OnlyInDB2 sheet has table_name and column_name columns
            db2_ws = wb['OnlyInDB2']
            db2_header = [cell.value for cell in db2_ws[1]]
            self.assertIn('table_name', db2_header)
            self.assertIn('column_name', db2_header)
            logging.info("test_compare_ddls passed and Excel file structure validated.")
        except Exception as e:
            logging.error(f"test_compare_ddls failed: {e}")
            self.fail(f"Exception occurred: {e}")

    def test_excel_writer(self):
        try:
            diff = pd.DataFrame({'a': [1]})
            only_db1 = pd.DataFrame({'b': [2]})
            only_db2 = pd.DataFrame({'c': [3]})
            ExcelWriter.write(diff, only_db1, only_db2, 'test_result.xlsx')
            logging.info("test_excel_writer passed.")
        except Exception as e:
            logging.error(f"test_excel_writer failed: {e}")
            self.fail(f"Exception occurred: {e}")

if __name__ == "__main__":
    unittest.main()
