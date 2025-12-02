"""Excel Report Writer Module

This module provides functionality to export database comparison results to Excel format
with automatic formatting and visual highlighting of differences.

The ExcelWriter class creates multi-sheet Excel workbooks with:
- DiffColumns: Columns with differences, with changed values highlighted in yellow
- OnlyInDB1: Columns unique to the primary database
- OnlyInDB2: Columns unique to the secondary database

Typical usage example:
    from src.excel.excel_writer import ExcelWriter
    import pandas as pd
    
    diff_df = pd.DataFrame({'col1': [1, 2]})
    only_db1_df = pd.DataFrame({'col2': [3, 4]})
    only_db2_df = pd.DataFrame({'col3': [5, 6]})
    
    ExcelWriter.write(diff_df, only_db1_df, only_db2_df, 'output.xlsx')
"""

import pandas as pd
import logging
from typing import Any


class ExcelWriter:
    """A class to write database comparison results to an Excel file with formatting.
    
    This static class provides methods to export comparison results to Excel format
    with automatic highlighting of differences for easy visual inspection.
    """
    
    @staticmethod
    def write(diff: pd.DataFrame, only_db1: pd.DataFrame, only_db2: pd.DataFrame, out_path: str) -> None:
        """Writes the comparison results to an Excel file with formatting.
        
        Creates a multi-sheet Excel workbook with automatic highlighting of differing columns.
        Columns ending with '_db1' or '_db2' in the DiffColumns sheet are highlighted in yellow.

        Args:
            diff (pd.DataFrame): A DataFrame containing the differences between the two databases.
                Expected to have columns with '_db1' and '_db2' suffixes for comparison.
            only_db1 (pd.DataFrame): A DataFrame containing columns only in the primary database.
            only_db2 (pd.DataFrame): A DataFrame containing columns only in the secondary database.
            out_path (str): The path to the output Excel file. Will be created or overwritten.
            
        Raises:
            PermissionError: If the output file is open or write-protected
            OSError: If the output path is invalid or disk is full
            Exception: If Excel writing or formatting fails
            
        Example:
            >>> ExcelWriter.write(diff_df, only_db1_df, only_db2_df, 'results.xlsx')
            >>> print("Results written successfully")
        """
        import openpyxl
        from openpyxl.styles import PatternFill
        logger = logging.getLogger(__name__)
        try:
            logger.info(f"Writing results to Excel file: {out_path}")
            with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
                diff.to_excel(writer, sheet_name='DiffColumns', index=False)
                only_db1.to_excel(writer, sheet_name='OnlyInDB1', index=False)
                only_db2.to_excel(writer, sheet_name='OnlyInDB2', index=False)
            # Highlight differences in DiffColumns sheet
            wb = openpyxl.load_workbook(out_path)
            ws = wb['DiffColumns']
            # Find columns that differ and highlight them
            header = [cell.value for cell in ws[1]]
            # Only process if header exists and has values
            if header and any(header):
                diff_cols = [i for i, col in enumerate(header) if col and (col.endswith('_db1') or col.endswith('_db2'))]
                fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                # Only iterate if there are data rows
                if ws.max_row and ws.max_row > 1:
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        for i in diff_cols:
                            if i < len(row) and row[i].value is not None:
                                row[i].fill = fill
            wb.save(out_path)
            logger.info("Excel file written and differences highlighted successfully.")
        except Exception as e:
            logger.error(f"Error writing Excel file: {e}")
            raise
